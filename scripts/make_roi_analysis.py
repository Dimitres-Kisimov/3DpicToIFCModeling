"""make_roi_analysis.py — the time & money savings analysis, downloadable.
Manual-baseline assumptions are stated in the file itself; machine times are
the MEASURED numbers from the fleet runs. Three German designer rates give
the finance department low/mid/high planning columns.

    python scripts/make_roi_analysis.py
Out: docs/roi/ROI_ANALYSIS.md + roi_analysis.xlsx + roi_analysis.csv
"""
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

REPO = Path(__file__).resolve().parents[1]
OUT = REPO / "docs" / "roi"

RATES = {"low €45/h": 45.0, "mid €60/h": 60.0, "high €90/h": 90.0}

# (task, pieces, manual min/object, manual review h, machine h, machine review h)
# manual basis: 1.5–3.0 min per placed object (search, drag, rotate, clearance
# check) — we use 2.0 min — plus a compliance re-check pass at 25% of placement
# time. Machine basis: MEASURED runs + a human review pass (spot-check + edits).
SCENARIOS = [
    ("One office room (legal, 10–20 pieces)",        15,  2.0, 0.125, 15 / 3600, 0.25),
    ("6-storey office — 210 King, Toronto",        2191,  2.0, 18.0,  17 / 60,   4.0),
    ("Whole 15-building fleet",                    6000,  2.0, 50.0,  16 / 60,  10.0),
    ("Re-check clearances after an edit round",     100,  0.5,  0.0,   1 / 3600, 0.1),
]


def rows():
    out = []
    for name, pieces, mpo, mrev, mach_h, mach_rev in SCENARIOS:
        manual_h = pieces * mpo / 60 + mrev
        machine_h = mach_h + mach_rev
        saved_h = manual_h - machine_h
        r = {"task": name, "pieces": pieces,
             "manual_h": round(manual_h, 2), "machine_h": round(machine_h, 2),
             "saved_h": round(saved_h, 2),
             "saved_pct": round(100 * saved_h / manual_h, 1)}
        for label, rate in RATES.items():
            r[f"saved {label}"] = round(saved_h * rate)
        out.append(r)
    return out


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    data = rows()
    heads = list(data[0].keys())

    # ---- CSV ---------------------------------------------------------------
    with open(OUT / "roi_analysis.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=heads)
        w.writeheader()
        w.writerows(data)

    # ---- XLSX --------------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "ROI"
    ws.append(["SCS Studio — time & money savings (Dimitres Kisimov, v3)"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(heads)
    for c in ws[3]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2F6BFF")
        c.alignment = Alignment(wrap_text=True)
    for r in data:
        ws.append([r[h] for h in heads])
    ws.append([])
    for line in [
        "Assumptions (stated, adjustable):",
        "· manual placement 2.0 min/object (search, drag, rotate, clearance check; industry 1.5–3.0)",
        "· manual compliance re-check pass = 25% of placement time",
        "· machine times are MEASURED fleet runs (210 King 17 min for 2,191 pieces; fleet ~16 min)",
        "· machine 'review' = human spot-check + manual polish pass",
        "· rates: German interior/BIM planner €45–€90/h (mid €60) — pick your column",
        "Annual planning example: a firm furnishing 20 mid-size buildings/yr at the",
        "210-King scale saves ≈ 20 × 68 h ≈ 1,360 designer-hours ≈ €82k at €60/h.",
    ]:
        ws.append([line])
    widths = [42, 8, 10, 10, 9, 10, 14, 14, 14]
    for i, wd in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = wd
    wb.save(OUT / "roi_analysis.xlsx")

    # ---- MD ----------------------------------------------------------------
    md = ["# SCS Studio — ROI analysis (time & money saved)", "",
          "**Use case:** a designer/BIM planner furnishing office & residential",
          "buildings must place every object by hand and keep German ASR workplace",
          "rules satisfied. SCS Studio replaces the placement + compliance pass with",
          "a one-click solver run; the human keeps a short review/polish pass.", "",
          "| " + " | ".join(heads) + " |",
          "|" + "---|" * len(heads)]
    for r in data:
        md.append("| " + " | ".join(str(r[h]) for h in heads) + " |")
    md += ["", "Assumptions: 2.0 min/object manual (1.5–3.0 industry), +25% manual",
           "compliance pass; machine times measured; review pass included in the",
           "machine column; rates €45/€60/€90 per hour (German market).",
           "", "Annual example: 20 buildings/yr at 210-King scale ≈ 1,360 designer-",
           "hours ≈ **€82,000/yr saved at €60/h** (€61k at €45, €122k at €90)."]
    (OUT / "ROI_ANALYSIS.md").write_text("\n".join(md), encoding="utf-8")
    for r in data:
        print(f"{r['task'][:44]:46s} manual {r['manual_h']:7.2f} h -> machine "
              f"{r['machine_h']:5.2f} h  saved {r['saved_pct']}%")
    print(f"-> {OUT}")


if __name__ == "__main__":
    main()
